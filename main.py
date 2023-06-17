import openpyxl
from openpyxl.styles import Font, Border, Side

def criar_tabela():
    atividades = []
    dias_semana = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]

    while True:
        atividade = input("Qual atividade deseja adicionar? (Digite 'fim' para encerrar): ")
        if atividade.lower() == "fim":
            break

        if atividade not in atividades:
            atividades.append(atividade)
            print("Atividade adicionada à lista.")
        else:
            print("Essa atividade já está presente na lista.")

    # Cria um novo arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define o estilo para os títulos e dias da semana
    titulo_font = Font(bold=True)
    dia_semana_font = Font(bold=True)

    # Define o estilo para as bordas da tabela
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Preenche os títulos da tabela com formatação
    sheet.cell(row=1, column=1, value="Atividade").font = titulo_font
    sheet.cell(row=1, column=1).border = border

    for col, dia in enumerate(dias_semana, start=2):
        sheet.cell(row=1, column=col, value=dia).font = titulo_font
        sheet.cell(row=1, column=col).border = border
        sheet.cell(row=1, column=col).border = Border(bottom=Side(border_style="thin"))

    # Preenche as atividades na tabela com formatação
    for row, atividade in enumerate(atividades, start=2):
        sheet.cell(row=row, column=1, value=atividade).border = border

        for col in range(2, len(dias_semana) + 2):
            sheet.cell(row=row, column=col).border = border

    # Salva o arquivo Excel
    workbook.save("tabela_atividades.xlsx")

# Chamada da função para criar a tabela
criar_tabela()
